#!/usr/bin/env python
# -*- coding: utf-8 -*-	
# Written by Yuen-Hsien Tseng on 2019/02/23
#   python joke2xml.py ../jokes_3691.xlsx > jokes_3691.xml
# Run next line on 2019/03/21
#   python joke2xml.py ../jokes_3414.xlsx > jokes_3414.xml
# On 2019/06/06
#   python joke2xml.py ../../joke_corpora/jokes_20190606.xlsx > joke_20190606.xml
# Modified from joke2xml.py by 許瑋倫 on 2019/06/21 to extract jokes from Excel to DB
#   python jokeExcel2DB.py 笑話標記.xlsx
import sys
import sqlite3
f = sys.argv[1]
from openpyxl import load_workbook
'''
scn2sen = { # list the worksheet (source) name mapping
    '笑話多多':'xhdd', # the value is the acronym of Hanyu Pinyin of the key
    '陳國棟': 'GD',
    '笑話大全':'xhdq',
    '笑話精選集':'xhjxj',
    '笑話集中營':'xhjzy',
    '小蕃薯':'xfs',
    'PTT':'PTT',
    'Dcard':'Dcard', # for some obvious website, the value is its key
    '巴哈姆特':'bhmt',
    '深藍論壇':'sllt',
    '國語日報':'gyrb',
    'Facebook':'Facebook',
    'Instagram':'Instagram',
    'Youtube':'Youtube',
    'Yahoo知識+':'Yahoo',
    '痞客邦':'pkb',
    '隨意窩':'syw',
    '名言佳句大全':'myjjdq',
    '壹讀':'yd',
    '每日頭條':'mrtt',
    '廁所文學':'cswx',
    '流行生活百業通':'lxshbyt',
    '消防資料庫':'xfzlk',
    '(網站)笑話集':'xhj',
    '笑話集錦':'xhjj',
    'Short-Funny.com':'ShortFunny',
    '人際分享':'rjfx',
    '其他網站':'qtwz',
    '歐美幽默笑話':'omymxh',
    '幽默宅急便':'ymzjb',
    '保證有笑〈火辣版〉':'bzyx',
    '幽默好有魅力':'ymhyml',
    '網路哈哈大笑話-趣聞軼事篇':'wlhhdxcw',
    '網路哈哈大笑話-三鮮笑湯':'wlhhdxsx',
    '誰會比我糗':'shbwq',
    '幽默，酷一點也好':'hmcydyh',
    '幽默‧諷刺‧笑話':'ymfcxh',
    '幽默的力量':'ymdll',
    '幽默開懷文選':'ymkhwx',
    '(APP)笑話集':'APPxhj',
    '老笑話':'lxh',
    '笑話連篇之笑死朕':'xhlpzxsz'
}
'''
wb = load_workbook(filename = f)
#print("Number of worksheets: {}".format(len(wb.sheetnames)))
for nwb, ws in enumerate(wb):
#    print(ws.title); continue
    for i, r in enumerate(ws.rows):
        if i == 0: # field name line
            field_names = r # list of field names
            continue
        if r[0].value is None: continue
        #pk = scn2sen[ws.title] + '_' + str(r[0].value)
        conn = sqlite3.connect('Chat_DB.db') #連結的資料庫名稱
        cur = conn.cursor()
        cur.execute("INSERT INTO TjokeBase (jokeID, title, txt, jokeScore) VALUES (?,?,?,?)"
            ,(str(r[0].value), str(r[1].value), r[2].value.strip(), str(r[3].value))) #存進的欄位和值
        conn.commit()
#        if i>2: break
#    if nwb>1: break